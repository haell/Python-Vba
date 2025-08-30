from pathlib import Path
from datetime import datetime
import csv
import openpyxl


# ==== CONFIGURACAO ====
ARQUIVOS_POSSIVEIS = ["tarefas.xlsm", "tarefas.xlsx"]
NOME_ABA = "Tarefas"


# ==== ENCONTRAR PLANILHA ====
wb_path = None
for nome in ARQUIVOS_POSSIVEIS:
	p = Path(nome)
	if p.exists():
		wb_path = p
		break
if wb_path is None:
	raise FileNotFoundError("Nenhum arquivo 'tarefas.xlsm' ou 'tarefas.xlsx' encontrado na pasta.")


# ==== CARREGAR EXCEL ====
wb = openpyxl.load_workbook(wb_path, data_only=True)
if NOME_ABA not in wb.sheetnames:
	raise ValueError(f"A aba '{NOME_ABA}' não foi encontrada.")
ws = wb[NOME_ABA]


# ==== LER DADOS (A:ID, B:Tarefa, C:Status, D:Criada_em, E:Concluida_em) ====
linhas = []
for row in ws.iter_rows(min_row=2, values_only=True):
	if row[0] is None and row[1] is None:
		continue # pula linhas vazias
	linhas.append({
		"ID": row[0],
		"Tarefa": row[1],
		"Status": row[2],
		"Criada_em": row[3],
		"Concluida_em": row[4],
	})


# ==== METRICAS ====
total = len(linhas)
pendentes = sum(1 for r in linhas if (r["Status"] or "").lower().startswith("pend"))
concluidas = sum(1 for r in linhas if (r["Status"] or "").lower().startswith("concl"))


# ==== SAIDAS ====
carimbo = datetime.now().strftime("%Y%m%d_%H%M%S")
Path("relatorios").mkdir(exist_ok=True)


# 1) CSV completo
csv_path = Path("relatorios") / f"tarefas_{carimbo}.csv"
with csv_path.open("w", newline="", encoding="utf-8") as f:
	w = csv.writer(f)
	w.writerow(["ID", "Tarefa", "Status", "Criada_em", "Concluida_em"])
	for r in linhas:
		w.writerow([r["ID"], r["Tarefa"], r["Status"], r["Criada_em"], r["Concluida_em"]])


# 2) Resumo TXT
resumo_path = Path("relatorios") / f"resumo_{carimbo}.txt"
with resumo_path.open("w", encoding="utf-8") as f:
	f.write("RELATORIO DE TAREFAS\n")
	f.write(f"Arquivo origem: {wb_path.name}\n")
	f.write(f"Total: {total}\n")
	f.write(f"Pendentes: {pendentes}\n")
	f.write(f"Concluidas: {concluidas}\n\n")
	f.write("Pendentes:\n")
	for r in linhas:
		if (r["Status"] or "").lower().startswith("pend"):
			f.write(f" - ({r['ID']}) {r['Tarefa']}\n")


print("Relatorios gerados:")
print(" - ", csv_path)
print(" - ", resumo_path)